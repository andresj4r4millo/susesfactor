from selenium import webdriver
import time
from selenium.webdriver.common.action_chains import ActionChains
#para el libro de exporte
import openpyxl
from unidecode import unidecode
import re
import pandas as pd
import numpy as np
from lxml import html
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
import traceback
from datetime import datetime
from dateutil.relativedelta import relativedelta
from selenium.webdriver.edge.service import Service as EdgeService
import os
from datetime import datetime
from tkinter import *
from tkinter import ttk
import tkinter as tk
def formatof(fecha):
    fechad = fecha.split("/")
    fechan = "".join(fechad)
    formato=fechan.split(" ")
    lafecha=formato[0]
    fechas=lafecha.split("-")
    fecha_n="".join(fechas)
    return fecha_n
###
def calcular_fechas():
    # Obtener la fecha actual
    fecha_actual = datetime.now().date()

    # Calcular la fecha de dos meses en adelante
    fecha_futura = fecha_actual + relativedelta(months=2)

    return  fecha_futura

# Llamar a la función y obtener las fechas
fecha_futura = calcular_fechas()
nuevo_workbook = openpyxl.Workbook()
nueva_sheet = nuevo_workbook.active

workbook = openpyxl.load_workbook('SSFF.xlsx', read_only=True, data_only=True, keep_links=False, keep_vba=False)
# Seleccionar la hoja de cálculo que deseas leer
sheet = workbook['Hoja1']
for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
    if index==1:
        continue
    cedula=str(row[0])
    ex=str(row[1])
    fecha=str(row[2])
    pais=str(row[3])
    nombre=str(row[4])
    apellido=str(row[5])
    campaña=str(row[6])
    semilla=str(row[7])
    celular=str(row[8])
    correo=str(row[9])
    correo_corporativo=str(row[10])
    direccion=str(row[11])
    barrio=str(row[12])
    preingreso=str(row[13])
    ingreso=str(row[14])
    rm=str(row[15])
    lider_virtual=str(row[16])
    estado=str(row[17])
    observaciones=str(row[18])
    fechan=formatof(fecha)
    fecha_ex=formatof(ex)

    sexo="FEMENINO"
    estadoc="CASADO"
    time.sleep(2)

    fecha_ft=fecha_futura.strftime('%d%m%Y')
    #//*[@id="main--globalSFHeader"]
    texto="Añadir trabajador temporal"

    #ssff()
    time.sleep(1)
    try:
        fechanaci=datetime.strptime(fechan,"%Y%m%d")
        fx= datetime.strptime(fecha_ex, "%Y%m%d")
    except ValueError:
        continue

        # Formatea la fecha en el nuevo formato (DDMMYYYY)
    fecha_expedicion = fx.strftime("%d%m%Y")
    fecha_naci=fechanaci.strftime("%d%m%Y")
    # Calcula la diferencia en años
    # Calcula la diferencia en años
    # Calcula la diferencia en días
    diferencia_en_dias = (fx - fechanaci).days

    # Calcula la diferencia en años
    diferencia_en_anios = diferencia_en_dias / 365

    # Comprueba si la diferencia es mayor o igual a 18 años
    if diferencia_en_anios >= 18:
        print("Han pasado más de 18 años entre la fecha de nacimiento y la fecha de expedición.")
    else:
        print("No han pasado 18 años entre la fecha de nacimiento y la fecha de expedición.")



    """
    nueva_sheet.cell(row=index, column=1, value=cedula)
    nueva_sheet.cell(row=index, column=2, value=pais)
    
    if fecha == None:
        continue
    else:
        fechan=formatof(fecha)
        expedicion=formatof(fecha_ex)
        # Convierte la fecha a un objeto datetime
        try:
            fecha_objeto = datetime.strptime(fechan, "%Y%m%d")
        except ValueError:
            continue

        # Formatea la fecha en el nuevo formato (DDMMYYYY)
        fecha_formateada = fecha_objeto.strftime("%d%m%Y")

        # Formatea la fecha en el nuevo formato (DD/MM/AAAA)
        fecha_formateada = fecha_objeto.strftime("%d%m%Y")
        print("nacimiento")
        print(fecha)
        print("expedicion")
        #print(fecha_ex)
        print(fecha_formateada)
        print(expedicion)
        #print(fechan)
    """
#nuevo_workbook.save('Nuevo_SSF.xlsx')

    



